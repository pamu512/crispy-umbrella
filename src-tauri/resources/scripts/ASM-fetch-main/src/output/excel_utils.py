"""
Excel writing utilities for domain scanner output.
"""

import logging
from typing import List, Dict

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from src.utils.formatters import format_opened_ports, format_whois, parse_possible_dict


logger = logging.getLogger(__name__)


def build_workbook(data: List[Dict[str, str]], headers: List[str]):
    """Build a workbook with multiple sheets (All In One, Subdomains, ASN, WHOIS, CVE, SPF, DMARC,
    DKIM, TLS SSL, Opened Ports, Unusual Ports, Sensitive Subdomains) and return it.
    """
    wb = openpyxl.Workbook()

    # Helper functions
    def set_header_bold_and_freeze(ws):
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
        ws.freeze_panes = ws["A2"]

    def autofit_worksheet_columns(ws):
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass
            adjusted_width = max(max_length, 5)
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    columns = headers

    # All In One sheet
    ws_all = wb.active
    ws_all.title = "All In One"
    ws_all.append(columns)
    for row in data:
        values = []
        for col in columns:
            if col == "Opened Ports":
                values.append(format_opened_ports(row.get(col, row.get('opened_ports', 'N/A'))))
            elif col == "WHOIS":
                v = row.get(col, "N/A")
                parsed = parse_possible_dict(v)
                if parsed is not None:
                    values.append(format_whois(parsed))
                else:
                    values.append(v if v is not None else "N/A")
            else:
                values.append(row.get(col, "N/A"))
        ws_all.append(values)
    set_header_bold_and_freeze(ws_all)
    autofit_worksheet_columns(ws_all)

    # Subdomains sheet (exclude rows with IPs == "N/A")
    ws_sub = wb.create_sheet("Subdomains")
    ws_sub.append(["Hosts", "IPs", "Type"]) 
    for row in data:
        if row.get("IPs") and row.get("IPs") != "N/A":
            ws_sub.append([row.get("Hosts", "N/A"), row.get("IPs", "N/A"), row.get("Type", "N/A")])
    set_header_bold_and_freeze(ws_sub)
    autofit_worksheet_columns(ws_sub)

    # ASN & ASN Name sheet
    ws_asn = wb.create_sheet("ASN & ASN Name")
    ws_asn.append(["Hosts", "IPs", "ASN", "ASN Name"])
    for row in data:
        if row.get("ASN") and row.get("ASN") != "N/A" and row.get("ASN Name") and row.get("ASN Name") != "N/A":
            ws_asn.append([row.get("Hosts", "N/A"), row.get("IPs", "N/A"), row.get("ASN", "N/A"), row.get("ASN Name", "N/A")])
    set_header_bold_and_freeze(ws_asn)
    autofit_worksheet_columns(ws_asn)

    # WHOIS sheet
    ws_whois = wb.create_sheet("WHOIS")
    ws_whois.append(["Hosts", "IPs", "WHOIS"])
    for row in data:
        v = row.get("WHOIS")
        formatted = None
        parsed = parse_possible_dict(v)
        if parsed is not None:
            formatted = format_whois(parsed)
        elif isinstance(v, str):
            formatted = v

        if formatted and formatted != "N/A":
            ws_whois.append([row.get("Hosts", "N/A"), row.get("IPs", "N/A"), formatted])
    set_header_bold_and_freeze(ws_whois)
    autofit_worksheet_columns(ws_whois)

    # CVE sheet
    ws_cve = wb.create_sheet("CVE")
    ws_cve.append(["Hosts", "IPs", "CVE"])
    for row in data:
        if row.get("CVE") and row.get("CVE") != "N/A":
            ws_cve.append([row.get("Hosts", "N/A"), row.get("IPs", "N/A"), row.get("CVE", "N/A")])
    set_header_bold_and_freeze(ws_cve)
    autofit_worksheet_columns(ws_cve)

    # SPF sheet
    ws_spf = wb.create_sheet("SPF")
    ws_spf.append(["Hosts", "IPs", "SPF"])
    for row in data:
        if row.get("SPF") and row.get("SPF") != "N/A":
            ws_spf.append([row.get("Hosts", "N/A"), row.get("IPs", "N/A"), row.get("SPF", "N/A")])
    set_header_bold_and_freeze(ws_spf)
    autofit_worksheet_columns(ws_spf)

    # DMARC sheet
    ws_dmarc = wb.create_sheet("DMARC")
    ws_dmarc.append(["Hosts", "IPs", "DMARC"])
    for row in data:
        if row.get("DMARC") and row.get("DMARC") != "N/A":
            ws_dmarc.append([row.get("Hosts", "N/A"), row.get("IPs", "N/A"), row.get("DMARC", "N/A")])
    set_header_bold_and_freeze(ws_dmarc)
    autofit_worksheet_columns(ws_dmarc)

    # DKIM sheet
    ws_dkim = wb.create_sheet("DKIM")
    ws_dkim.append(["Hosts", "IPs", "DKIM"])
    for row in data:
        if row.get("DKIM") and row.get("DKIM") != "N/A":
            ws_dkim.append([row.get("Hosts", "N/A"), row.get("IPs", "N/A"), row.get("DKIM", "N/A")])
    set_header_bold_and_freeze(ws_dkim)
    autofit_worksheet_columns(ws_dkim)

    # TLS SSL sheet
    ws_tls = wb.create_sheet("TLS SSL")
    ws_tls.append(["Hosts", "IPs", "TLS SSL"])
    for row in data:
        if row.get("TLS SSL") and row.get("TLS SSL") != "N/A":
            ws_tls.append([row.get("Hosts", "N/A"), row.get("IPs", "N/A"), row.get("TLS SSL", "N/A")])
    set_header_bold_and_freeze(ws_tls)
    autofit_worksheet_columns(ws_tls)

    # Opened Ports
    ws_ports = wb.create_sheet("Opened Ports")
    ws_ports.append(["Hosts", "IPs", "Opened Ports"]) 
    for row in data:
        if row.get("Opened Ports") and row.get("Opened Ports") != "N/A":
            ws_ports.append([
                row.get("Hosts", "N/A"),
                row.get("IPs", "N/A"),
                format_opened_ports(row.get("Opened Ports", row.get('opened_ports', 'N/A')))
            ])
    set_header_bold_and_freeze(ws_ports)
    autofit_worksheet_columns(ws_ports)

    # Unusual Ports
    ws_unusual = wb.create_sheet("Unusual Ports")
    ws_unusual.append(["Hosts", "IPs", "Unusual Ports"]) 
    for row in data:
        if row.get("Unusual Ports") and row.get("Unusual Ports") != "N/A":
            ws_unusual.append([row.get("Hosts", "N/A"), row.get("IPs", "N/A"), row.get("Unusual Ports", "N/A")])
    set_header_bold_and_freeze(ws_unusual)
    autofit_worksheet_columns(ws_unusual)

    # Sensitive Subdomains (only three columns)
    sensitive_columns = ["Hosts", "IPs", "Sensitive Subdomains"]
    sensitive_rows = [row for row in data if row.get("Sensitive Subdomains") and row.get("Sensitive Subdomains") != "N/A"]
    ws_sensitive = wb.create_sheet("Sensitive Subdomains")
    ws_sensitive.append(sensitive_columns)
    for row in sensitive_rows:
        ws_sensitive.append([row.get("Hosts", "N/A"), row.get("IPs", "N/A"), row.get("Sensitive Subdomains", "N/A")])
    set_header_bold_and_freeze(ws_sensitive)
    autofit_worksheet_columns(ws_sensitive)

    logger.info(f"Added Sensitive Subdomains sheet with {len(sensitive_rows)} rows")

    return wb