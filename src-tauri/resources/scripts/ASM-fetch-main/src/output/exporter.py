"""
Data export module for CSV and XLSX.
"""

import csv
import logging
from io import BytesIO
from typing import List, Dict
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None
from src.utils.formatters import extract_unusual_ports  # Not used here, but kept for consistency
from src.utils.formatters import format_opened_ports, parse_possible_dict
from src.utils.formatters import format_whois


logger = logging.getLogger(__name__)


def export_data(csv_data: List[Dict], output_base: str, output_formats: str, verbose: bool = False) -> None:
    """
    Export data to CSV and/or XLSX formats.

    Args:
        csv_data: List of row dicts.
        output_base: Base filename (e.g., 'blackwired_com').
        output_formats: Comma-separated formats (csv, xlsx).
        verbose: Verbose mode.
    """
    formats = [fmt.strip().lower() for fmt in output_formats.split(",") if fmt.strip().lower() in ("csv", "xlsx")]
    if not formats:
        formats = ["xlsx"]

    output_csv = f'{output_base}_subdomains.csv'
    output_xlsx = f'{output_base}_subdomains.xlsx'

    separator_printed = False
    for fmt in formats:
        if fmt == "csv":
            _export_csv(csv_data, output_csv)
            if not separator_printed:
                print("------------------------------------------------------------------------------")
                separator_printed = True
            logger.info(f"Saved {len(csv_data)} entries to {output_csv}")
        elif fmt == "xlsx":
            _export_xlsx(csv_data, output_xlsx)
            if not separator_printed:
                print("------------------------------------------------------------------------------")
                separator_printed = True
            logger.info(f"Saved {len(csv_data)} entries to {output_xlsx}")


def _export_csv(csv_data: List[Dict], output_csv: str) -> None:
    """Export to CSV with all columns, including Sensitive Subdomains and TLS SSL."""
    columns = [
        "Hosts", "IPs", "Type", "ASN", "ASN Name", "WHOIS", "CVE", "SPF", "DMARC", "DKIM",
        "TLS SSL", "Opened Ports", "Unusual Ports", "Sensitive Subdomains"
    ]
    # Ensure complex fields are serialized to strings to avoid CSV column shifting
    def _normalize_row(row: Dict) -> Dict:
        out = {}
        import ast
        for col in columns:
            val = row.get(col, "N/A")
            if col == "WHOIS":
                parsed = parse_possible_dict(val)
                out[col] = format_whois(parsed)
            elif isinstance(val, dict):
                # special-case tls_ssl: serialize dict back to legacy key:value semicolon/newline format
                if col == "TLS SSL":
                    try:
                        parts = []
                        for k, v in val.items():
                            if isinstance(v, list):
                                parts.append(f"{k}:{'; '.join(map(str,v))}")
                            else:
                                parts.append(f"{k}:{v}")
                        out[col] = "; ".join(parts) if parts else "N/A"
                    except Exception:
                        out[col] = str(val)
                else:
                    try:
                        out[col] = str(val) if val else "N/A"
                    except Exception:
                        out[col] = "N/A"
            elif isinstance(val, list):
                try:
                    out[col] = "; ".join(str(x) for x in val) if val else "N/A"
                except Exception:
                    out[col] = "N/A"
            else:
                out[col] = val if val is not None else "N/A"
        # Make sure Unusual Ports is computed from Opened Ports if it's a list
        opened = row.get("Opened Ports", row.get("opened_ports", []))
        out["Opened Ports"] = format_opened_ports(opened)
        # If Unusual Ports is a list, stringify it
        up = row.get("Unusual Ports", row.get("unusual_ports", "N/A"))
        if isinstance(up, list):
            out["Unusual Ports"] = ",".join(str(x) for x in up) if up else "N/A"
        # Sensitive Subdomains ensure string
        ss = row.get("Sensitive Subdomains", row.get("sensitive_subdomains", "N/A"))
        if isinstance(ss, list):
            out["Sensitive Subdomains"] = ", ".join(ss) if ss else "N/A"
        else:
            out["Sensitive Subdomains"] = ss if ss else "N/A"
        # CVE: ensure list serialized as comma-separated string for exports
        cv = row.get("CVE", row.get("cve", "N/A"))
        if isinstance(cv, list):
            out["CVE"] = ",".join(cv) if cv else "N/A"
        else:
            # if CVE was already set earlier in the loop, keep it; otherwise set from val
            out["CVE"] = out.get("CVE") if out.get("CVE") is not None else (cv if cv else "N/A")
        return out

    with open(output_csv, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=columns)
        writer.writeheader()
        for row in csv_data:
            writer.writerow(_normalize_row(row))


def _export_xlsx(csv_data: List[Dict], output_xlsx: str) -> None:
    """Export to XLSX with multiple sheets, including Sensitive Subdomains."""
    if openpyxl is None:
        logger.error("openpyxl is not installed. Please run 'pip install openpyxl' to use XLSX output.")
        return

    wb = Workbook()
    columns = [
        "Hosts", "IPs", "Type", "ASN", "ASN Name", "WHOIS", "CVE", "SPF", "DMARC", "DKIM",
        "TLS SSL", "Opened Ports", "Unusual Ports", "Sensitive Subdomains"
    ]

    # All In One sheet
    ws_all = wb.active
    ws_all.title = "All In One"
    ws_all.append(columns)
    import ast
    for row in csv_data:
        values = []
        for col in columns:
            if col == "WHOIS":
                val = row.get(col, "N/A")
                if isinstance(val, str) and val.strip().startswith("{") and val.strip().endswith("}"):
                    try:
                        val = ast.literal_eval(val)
                    except Exception:
                        pass
                values.append(format_whois(val))
            elif col == "Opened Ports":
                values.append(format_opened_ports(row.get(col, row.get('opened_ports', 'N/A'))))
            elif col == "CVE":
                val = row.get(col, "N/A")
                if isinstance(val, list):
                    values.append(",".join(val) if val else "N/A")
                else:
                    values.append(val)
            else:
                values.append(row.get(col, "N/A"))
        ws_all.append(values)
    set_header_bold_and_freeze(ws_all)
    autofit_worksheet_columns(ws_all)

    # Subdomains sheet (exclude rows with IPs == "N/A")
    ws_sub = wb.create_sheet("Subdomains")
    ws_sub.append(["Hosts", "IPs", "Type"])
    for row in csv_data:
        if row["IPs"] != "N/A":
            ws_sub.append([row["Hosts"], row["IPs"], row["Type"]])
    set_header_bold_and_freeze(ws_sub)
    autofit_worksheet_columns(ws_sub)

    # ASN & ASN Name sheet (exclude rows with ASN == "N/A" and ASN Name == "N/A")
    ws_asn = wb.create_sheet("ASN & ASN Name")
    ws_asn.append(["Hosts", "IPs", "ASN", "ASN Name"])
    for row in csv_data:
        if row["ASN"] != "N/A" and row["ASN Name"] != "N/A":
            ws_asn.append([row["Hosts"], row["IPs"], row["ASN"], row["ASN Name"]])
    set_header_bold_and_freeze(ws_asn)
    autofit_worksheet_columns(ws_asn)

    # WHOIS sheet (exclude rows with WHOIS == "N/A")
    ws_whois = wb.create_sheet("WHOIS")
    ws_whois.append(["Hosts", "IPs", "WHOIS"])
    import ast
    for row in csv_data:
        val = row.get("WHOIS", "N/A")
        if isinstance(val, str) and val.strip().startswith("{") and val.strip().endswith("}" ):
            try:
                val = ast.literal_eval(val)
            except Exception:
                pass
        formatted_whois = format_whois(val)
        if formatted_whois != "N/A":
            ws_whois.append([row["Hosts"], row["IPs"], formatted_whois])
    set_header_bold_and_freeze(ws_whois)
    autofit_worksheet_columns(ws_whois)

    # CVE sheet (exclude rows with CVE == "N/A")
    ws_cve = wb.create_sheet("CVE")
    ws_cve.append(["Hosts", "IPs", "CVE"])
    for row in csv_data:
        if row["CVE"] != "N/A":
            ws_cve.append([row["Hosts"], row["IPs"], row["CVE"]])
    set_header_bold_and_freeze(ws_cve)
    autofit_worksheet_columns(ws_cve)

    # SPF sheet (exclude rows with SPF == "N/A")
    ws_spf = wb.create_sheet("SPF")
    ws_spf.append(["Hosts", "IPs", "SPF"])
    for row in csv_data:
        if row["SPF"] != "N/A":
            ws_spf.append([row["Hosts"], row["IPs"], row["SPF"]])
    set_header_bold_and_freeze(ws_spf)
    autofit_worksheet_columns(ws_spf)

    # DMARC sheet (exclude rows with DMARC == "N/A")
    ws_dmarc = wb.create_sheet("DMARC")
    ws_dmarc.append(["Hosts", "IPs", "DMARC"])
    for row in csv_data:
        if row["DMARC"] != "N/A":
            ws_dmarc.append([row["Hosts"], row["IPs"], row["DMARC"]])
    set_header_bold_and_freeze(ws_dmarc)
    autofit_worksheet_columns(ws_dmarc)

    # DKIM sheet (exclude rows with DKIM == "N/A")
    ws_dkim = wb.create_sheet("DKIM")
    ws_dkim.append(["Hosts", "IPs", "DKIM"])
    for row in csv_data:
        if row["DKIM"] != "N/A":
            ws_dkim.append([row["Hosts"], row["IPs"], row["DKIM"]])
    set_header_bold_and_freeze(ws_dkim)
    autofit_worksheet_columns(ws_dkim)

    # TLS SSL sheet (always create, even if empty)
    ws_tls = wb.create_sheet("TLS SSL")
    ws_tls.append(["Hosts", "IPs", "TLS SSL"])
    for row in csv_data:
        if row["TLS SSL"] != "N/A":
            ws_tls.append([row["Hosts"], row["IPs"], row["TLS SSL"]])
    set_header_bold_and_freeze(ws_tls)
    autofit_worksheet_columns(ws_tls)

    # Opened Ports sheet (exclude rows with Opened Ports == "N/A")
    ws_ports = wb.create_sheet("Opened Ports")
    ws_ports.append(["Hosts", "IPs", "Opened Ports"])
    for row in csv_data:
        if row.get("Opened Ports") and row["Opened Ports"] != "N/A":
            ws_ports.append([row["Hosts"], row["IPs"], format_opened_ports(row.get("Opened Ports", row.get('opened_ports', 'N/A')))])
    set_header_bold_and_freeze(ws_ports)
    autofit_worksheet_columns(ws_ports)

    # Unusual Ports sheet (exclude rows with Unusual Ports == "N/A")
    ws_unusual = wb.create_sheet("Unusual Ports")
    ws_unusual.append(["Hosts", "IPs", "Unusual Ports"])
    for row in csv_data:
        if row.get("Unusual Ports") and row["Unusual Ports"] != "N/A":
            ws_unusual.append([row["Hosts"], row["IPs"], row["Unusual Ports"]])
    set_header_bold_and_freeze(ws_unusual)
    autofit_worksheet_columns(ws_unusual)

    # Sensitive Subdomains sheet (always create, even if empty)
    sensitive_columns = ["Hosts", "IPs", "Sensitive Subdomains"]
    sensitive_rows = [row for row in csv_data if row.get("Sensitive Subdomains") != "N/A"]
    ws_sensitive = wb.create_sheet("Sensitive Subdomains")
    ws_sensitive.append(sensitive_columns)
    for row in sensitive_rows:
        ws_sensitive.append([row.get(col, "N/A") for col in sensitive_columns])
    set_header_bold_and_freeze(ws_sensitive)
    autofit_worksheet_columns(ws_sensitive)
    logger.info(f"Added Sensitive Subdomains sheet with {len(sensitive_rows)} rows")

    wb.save(output_xlsx)


def set_header_bold_and_freeze(ws):
    """Set the first row bold and freeze it for the worksheet."""
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    ws.freeze_panes = ws["A2"]


def autofit_worksheet_columns(ws):
    """Auto-fit column widths in a worksheet."""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value else 0
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass
        adjusted_width = max(max_length, 5)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width